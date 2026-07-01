import csv
import io
from datetime import timedelta

from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Count, Q
from django.utils import timezone

from CRM.models import Lead, FollowUp, CRMUser, AuditLog
from CRM.decorators import crm_login_required, admin_required


@crm_login_required
def reports_index(request):
    """Reports dashboard with all report types."""
    user = CRMUser.objects.get(id=request.session['crm_user_id'])
    today = timezone.now().date()

    # Date filters
    date_from = request.GET.get('date_from', (today - timedelta(days=30)).isoformat())
    date_to = request.GET.get('date_to', today.isoformat())
    report_type = request.GET.get('report_type', 'lead')

    if user.is_admin:
        leads = Lead.objects.all()
    else:
        leads = Lead.objects.filter(assigned_to=user)

    if date_from:
        leads = leads.filter(created_at__date__gte=date_from)
    if date_to:
        leads = leads.filter(created_at__date__lte=date_to)

    # Lead Report
    lead_report = leads.select_related('assigned_to').order_by('-created_at')

    # Status Report
    status_report = list(
        leads.values('status').annotate(count=Count('id')).order_by('status')
    )

    # Category Report
    category_report = list(
        leads.values('category').annotate(count=Count('id')).order_by('category')
    )

    # User Performance Report
    user_performance = []
    if user.is_admin:
        all_users = CRMUser.objects.filter(is_active=True)
        for u in all_users:
            user_leads = leads.filter(assigned_to=u)
            user_performance.append({
                'user': u,
                'total': user_leads.count(),
                'new': user_leads.filter(status='new').count(),
                'closed': user_leads.filter(status='deal_closed').count(),
                'not_useful': user_leads.filter(status='not_useful').count(),
            })

    # Follow-Up Report
    if user.is_admin:
        followups = FollowUp.objects.all()
    else:
        followups = FollowUp.objects.filter(assigned_to=user)

    if date_from:
        followups = followups.filter(date__gte=date_from)
    if date_to:
        followups = followups.filter(date__lte=date_to)

    followup_report = {
        'total': followups.count(),
        'pending': followups.filter(status='pending').count(),
        'completed': followups.filter(status='completed').count(),
        'overdue': followups.filter(status='overdue').count(),
    }

    context = {
        'lead_report': lead_report[:100],
        'status_report': status_report,
        'category_report': category_report,
        'user_performance': user_performance,
        'followup_report': followup_report,
        'date_from': date_from,
        'date_to': date_to,
        'report_type': report_type,
        'total_leads': leads.count(),
        'status_choices': Lead.STATUS_CHOICES,
        'category_choices': Lead.CATEGORY_CHOICES,
    }
    return render(request, 'crm/reports/index.html', context)


@crm_login_required
def export_report(request):
    """Export report as CSV or Excel."""
    user = CRMUser.objects.get(id=request.session['crm_user_id'])
    export_type = request.GET.get('type', 'csv')
    report_type = request.GET.get('report_type', 'lead')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if user.is_admin:
        leads = Lead.objects.all()
    else:
        leads = Lead.objects.filter(assigned_to=user)

    if date_from:
        leads = leads.filter(created_at__date__gte=date_from)
    if date_to:
        leads = leads.filter(created_at__date__lte=date_to)

    leads = leads.select_related('assigned_to').order_by('-created_at')

    if export_type == 'excel':
        return _export_excel(leads, report_type)
    else:
        return _export_csv(leads, report_type)


def _export_csv(leads, report_type):
    """Export leads to CSV."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Name', 'Phone', 'Category', 'Status', 'Assigned To',
        'Source', 'Remarks', 'Follow-Up Date', 'Created Date'
    ])

    for lead in leads:
        writer.writerow([
            lead.id,
            lead.name,
            lead.phone,
            lead.get_category_display(),
            lead.get_status_display(),
            lead.assigned_to.name if lead.assigned_to else '',
            lead.get_source_display(),
            lead.remarks,
            lead.follow_up_date or '',
            lead.created_at.strftime('%Y-%m-%d %H:%M'),
        ])

    return response


def _export_excel(leads, report_type):
    """Export leads to Excel using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        return _export_csv(leads, report_type)

    wb = Workbook()
    ws = wb.active
    ws.title = f'{report_type.title()} Report'

    # Header styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

    headers = [
        'ID', 'Name', 'Phone', 'Category', 'Status', 'Assigned To',
        'Source', 'Remarks', 'Follow-Up Date', 'Created Date'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_num, lead in enumerate(leads, 2):
        ws.cell(row=row_num, column=1, value=lead.id)
        ws.cell(row=row_num, column=2, value=lead.name)
        ws.cell(row=row_num, column=3, value=lead.phone)
        ws.cell(row=row_num, column=4, value=lead.get_category_display())
        ws.cell(row=row_num, column=5, value=lead.get_status_display())
        ws.cell(row=row_num, column=6, value=lead.assigned_to.name if lead.assigned_to else '')
        ws.cell(row=row_num, column=7, value=lead.get_source_display())
        ws.cell(row=row_num, column=8, value=lead.remarks)
        ws.cell(row=row_num, column=9, value=str(lead.follow_up_date or ''))
        ws.cell(row=row_num, column=10, value=lead.created_at.strftime('%Y-%m-%d %H:%M'))

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report.xlsx"'
    return response
